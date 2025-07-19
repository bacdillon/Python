<<<<<<< HEAD
""" This script reads data from an Excel file (input_data.xlsx), 
checks the "Score" column, and highlights cells with values greater 
than 50. Highlighted cells are formatted with a Reflex Blue background and 
white font. The modified data is saved to a new file 
(output_highlighted.xlsx) using the openpyxl library. """

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# File paths
input_file = "input_data.xlsx"
output_file = "output_highlighted.xlsx"

# Load workbook and worksheet
wb = load_workbook(input_file)
ws = wb.active  # use wb["SheetName"] if needed

# Define yellow fill and white font
highlight_fill = PatternFill(start_color="001489", end_color="001489", fill_type="solid")
white_font = Font(color="FFFFFF")  # White text

# Apply formatting to Score column if > 50
for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):  # Column B
    for cell in row:
        if isinstance(cell.value, (int, float)) and cell.value > 50:
            cell.fill = highlight_fill
            cell.font = white_font

# Save the result
wb.save(output_file)
print(f"Highlighted cells with white text saved to: {output_file}")
=======
""" This script reads data from an Excel file (input_data.xlsx), 
checks the "Score" column, and highlights cells with values greater 
than 50. Highlighted cells are formatted with a Reflex Blue background and 
white font. The modified data is saved to a new file 
(output_highlighted.xlsx) using the openpyxl library. """

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# File paths
input_file = "input_data.xlsx"
output_file = "output_highlighted.xlsx"

# Load workbook and worksheet
wb = load_workbook(input_file)
ws = wb.active  # use wb["SheetName"] if needed

# Define yellow fill and white font
highlight_fill = PatternFill(start_color="001489", end_color="001489", fill_type="solid")
white_font = Font(color="FFFFFF")  # White text

# Apply formatting to Score column if > 50
for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):  # Column B
    for cell in row:
        if isinstance(cell.value, (int, float)) and cell.value > 50:
            cell.fill = highlight_fill
            cell.font = white_font

# Save the result
wb.save(output_file)
print(f"Highlighted cells with white text saved to: {output_file}")
>>>>>>> bf39258ab5c050e50de5ea7fd8308521162ac5f4
